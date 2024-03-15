
from customtkinter import*
from tkinter import filedialog
from openpyxl import load_workbook
import random
from PIL import Image,ImageTk
from trio import Path


uppercase=[char for char in range(65,91)]
lowercase=[char for char in range(97,123)]
numbers=[char for char in range(48,58)]
specialchar=[35,36,64]

class App:
    def __init__(self,root) -> None:
        self.root=root
        root.title("Generator")
        root.geometry("650x500")
        root.resizable(False,False)
        set_appearance_mode("light")
        set_default_color_theme("dark-blue")
        set_widget_scaling(1.2)
        
        ###### icon windows ######################################################################
        # self.iconpath=ImageTk.PhotoImage(file=os.path.join("assets","checked.ico"))
        # root.wm_iconbitmap()
        # root.iconphoto(False,iconpath)
        ####################################################################################
        self.label=CTkLabel(root,text="Style with 8 characters with (# $ @)").place(x=25,y=70)
        self.sv=StringVar()
        self.style =CTkEntry(root,textvariable = self.sv,width=250,height=30)
        self.style.place(x=20,y=100)
        self.sv.trace("w", lambda name, index, mode, sv=self.sv: self.callback(self.sv))
        self.label=CTkLabel(root,text="number of output style").place(x=330,y=70)
        self.stylenumber =CTkEntry(root ,width=130,height=30)
        self.stylenumber.place(x=330,y=100)

        # self.label=CTkLabel(root,text="number of output style").place(x=25,y=150)
        self.excellpath =CTkEntry(root ,width=360,height=30,placeholder_text="output Excell file example : file.xlsx")
        self.excellpath.place(x=20,y=180)
        self.button_choose =CTkButton(root ,width=30,height=30,text="choose",command=self.open_file_dialog_entry)
        self.button_choose.place(x=390,y=180)

        self.procbar=CTkProgressBar(master=root,width=400,height=10,mode="determinate")
        self.procbar.set(0)
        self.procbar.place(relx=0.17,rely=0.7)

        self.labelproc=CTkLabel(master=root,text="generating...")
        self.labelproc.place(relx=0.17,rely=0.61)

        self.button_generate =CTkButton(root ,width=30,height=30,text="generate",command=self.outs)
        self.button_generate.place(x=240,y=350)



    def callback(self,sv):
        return self.sv.set(self.sv.get()[:8])
    
    def open_file_dialog_entry(self):
        file_path = filedialog.askopenfilename(title="Select a file")
        if file_path:
            self.excellpath.delete(0,END)
            self.excellpath.insert(0,file_path)
            

            return file_path
            # root.destroy()
        else:
            return None
        
    def style_value(self)->str:
        return self.sv.get()

    def out_num_value(self)->str:
        return self.stylenumber.get()
    
    def excelle_Path(self)->str:
        return self.excellpath.get()


    def domains(self)->list:
        text=self.style_value()
        domain=[]
        for i in text:
            charcter_ascii=ord(i)
            if charcter_ascii in uppercase:
                val=uppercase
                domain.append(val)
                continue
            elif charcter_ascii in lowercase:

                val=lowercase
                domain.append(val)
                continue 
            elif charcter_ascii in numbers:
                val=numbers
                domain.append(val)
                continue
            elif charcter_ascii in specialchar:
                val=specialchar
                domain.append(val)
                continue
        
        return domain
    
    def generate(self):



        
        outset=set()
        need=int(self.out_num_value())
        
        n = len(outset)-1
        iter_step = 1/n
        r=(abs(int(iter_step)-(iter_step)))/iter_step
        
        progress_step = iter_step
        # progress_step = iter_step
        self.procbar.start()
        while len(outset) < need:
            password=f"{chr(random.choice(self.domains()[0]))}{chr(random.choice(self.domains()[1]))}{chr(random.choice(self.domains()[2]))}{chr(random.choice(self.domains()[3]))}{chr(random.choice(self.domains()[4]))}{chr(random.choice(self.domains()[5]))}{chr(random.choice(self.domains()[6]))}{chr(random.choice(self.domains()[7]))}"

            outset.add(password)

            #---------------------------------------------------------------------------------
            self.labelproc.configure(text=f"generating ....")
            self.labelproc.update()
            self.procbar.set(progress_step)
            progress_step += iter_step
            #-------------------------------------------------------------------------------
        self.procbar.set(0)
        return list(outset)  
    
   
    def clear_column_except_header(self,sheet, column_index):
        for row_num, row in enumerate(sheet.iter_rows(min_col=column_index, max_col=column_index), start=1):
            for cell in row:
                if row_num > 1:  # Skip the header row
                    cell.value = None  # You can also use an empty string by setting cell.value = ''

 
    
    def clear_garbage(self,output_path):

            
        wb = load_workbook(fr"{output_path}")
        sheet = wb["Sheet1"]

        
        for column_to_clear in range(1,2):  # Replace with the desired column index (1-based index)
            self.clear_column_except_header(sheet, column_to_clear)

        # Save the changes
        wb.save(fr"{output_path}")
        wb.close()

    def outs(self):
        path=self.excelle_Path()
        
        self.clear_garbage(path)
        
        self.procbar.stop()

        self.procbar.set(0)

        wb=load_workbook(fr"{path}")
        ws=wb["Sheet1"]
        

        
        
        # progress_step = iter_step
        self.procbar.start()
        passwordlist=self.generate()
        n = len(passwordlist)-1
        iter_step = 1/n
        r=(abs(int(iter_step)-(iter_step)))/iter_step
        progress_step = iter_step
        for i in range(1,len(passwordlist)+1):
            password=passwordlist[i-1]
            ws[f"A{i}"]=password

            #---------------------------------------------------------------------------------
            self.labelproc.configure(text=f"setting output ....")
            self.labelproc.update()
            self.procbar.set(progress_step)
            progress_step += iter_step
            #-------------------------------------------------------------------------------




        wb.save(fr"{path}")
        wb.close()

        self.procbar.stop()
        self.labelproc.configure(text=f"Done!!!! ....")
        self.labelproc.update()











root=CTk()

App(root)

root.mainloop()